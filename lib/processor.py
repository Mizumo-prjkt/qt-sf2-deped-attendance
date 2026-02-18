import openpyxl
from openpyxl.styles import PatternFill
import datetime
import calendar
import os

# Cell Coordinates
COORD_SCHOOL_ID = "G6"
COORD_SCHOOL_NAME = "G7"
COORD_SCHOOL_YEAR = "N6"
COORD_MONTH = "AA6"
COORD_GRADE = "AA7"
COORD_SECTION = "AF7"
COORD_TOTAL_DAYS = "AQ9" 

# Rows
ROW_HEADER_DATE = 10
ROW_HEADER_DAY = 11

ROW_START_MALE = 13
ROW_END_MALE = 42

ROW_START_FEMALE = 44
ROW_END_FEMALE = 73

ROW_DAILY_TOTAL_MALE = 43
ROW_DAILY_TOTAL_FEMALE = 74
ROW_DAILY_TOTAL_COMBINED = 75

# Columns (G=7 to AE=31)
COL_START_IDX = 7 
COL_END_IDX = 31

# Stats Columns
COL_ABSENT = 32 # AF
COL_PRESENT = 33 # AG

def get_weekdays_in_month(school_year_str, month_name):
    """
    Returns a list of datetime objects for all weekdays (Mon-Fri) in the given month.
    Handles School Year "2025-2026".
    Logic:
    - If Month is Jun-Dec: Use Start Year.
    - If Month is Jan-May: Use End Year.
    - If Jan: Start on first Monday >= Jan 4.
    """
    month_map = {name: num for num, name in enumerate(calendar.month_name) if num}
    month_num = month_map.get(month_name)
    
    if not month_num:
        try:
             month_num = int(month_name)
        except:
             raise ValueError(f"Invalid month name: {month_name}")

    try:
        if "-" in str(school_year_str):
            parts = str(school_year_str).split("-")
            start_year = int(parts[0])
            end_year = int(parts[1])
        else:
            start_year = int(school_year_str)
            end_year = start_year + 1
    except ValueError:
         raise ValueError(f"Invalid School Year format: {school_year_str}")

    if 1 <= month_num <= 5:
        target_year = end_year
    else:
        target_year = start_year
        
    num_days = calendar.monthrange(target_year, month_num)[1]
    weekdays = []
    
    for day in range(1, num_days + 1):
        dt = datetime.date(target_year, month_num, day)
        if dt.weekday() < 5: # 0=Mon, 4=Fri
            # January Logic: Skip if before Jan 4
            if month_num == 1 and day < 4:
                continue
            # If Jan 4 is reached, we must ensure we start on a Monday?
            # User said "Start on first Monday >= Jan 4".
            # If Jan 4 is Tue, we wait for next Mon? Or just start?
            # User example: "Jan 1 Mon -> Jan 8 start". (Jan 8 is Mon). 
            # "Jan 1 Thu -> Jan 5 start". (Jan 5 is Mon).
            # So basically, find first Monday >= Jan 4.
            if month_num == 1:
                 # Check if we have found the start date yet
                 if not weekdays:
                     if dt.weekday() != 0: # Not Monday
                         continue
            
            weekdays.append(dt.strftime("%Y-%m-%d"))
            
    return weekdays

def save_to_excel(data, template_path, output_path):
    """
    Saves data to SF2 Template with Formulas.
    data includes: 'holidays': set of date_strings
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")
        
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active 
    
    # 1. Fill Header Info
    sheet[COORD_SCHOOL_NAME] = data.get("school_name", "")
    sheet[COORD_SCHOOL_ID] = data.get("school_id", "")
    sheet[COORD_SCHOOL_YEAR] = data.get("school_year", "")
    sheet[COORD_MONTH] = data.get("month", "")
    sheet[COORD_GRADE] = data.get("grade", "")
    sheet[COORD_SECTION] = data.get("section", "")
    
    # 2. Setup Dates & Holidays
    dates = data.get("dates", [])
    holidays = data.get("holidays", set())
    
    current_col = COL_START_IDX
    date_col_map = {}
    valid_days_count = 0
    
    for date_str in dates:
        if current_col > COL_END_IDX:
            break 
            
        date_col_map[date_str] = current_col
        col_letter = openpyxl.utils.get_column_letter(current_col)
        
        # Write Day number to Row 10 ONLY if NOT holiday
        is_holiday = date_str in holidays
        if not is_holiday:
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            sheet.cell(row=ROW_HEADER_DATE, column=current_col, value=dt.day)
            valid_days_count += 1
        else:
            # Clear it just in case template has something
            sheet.cell(row=ROW_HEADER_DATE, column=current_col, value="")

        # Inject Daily Total Formulas (Rows 43, 74, 75)
        # Male Total: Row 43
        # =IF(G10="","",COUNTA($B$13:$B$42)-(COUNTIF(G13:G42,"x") + COUNTIF(G13:G42,"h")*0.5))
        f_male = f'=IF({col_letter}10="","",COUNTA($B$13:$B$42)-(COUNTIF({col_letter}13:{col_letter}42,"x") + COUNTIF({col_letter}13:{col_letter}42,"h")*0.5))'
        sheet.cell(row=ROW_DAILY_TOTAL_MALE, column=current_col, value=f_male)
        
        # Female Total: Row 74
        f_female = f'=IF({col_letter}10="","",COUNTA($B$44:$B$73)-(COUNTIF({col_letter}44:{col_letter}73,"x") + COUNTIF({col_letter}44:{col_letter}73,"h")*0.5))'
        sheet.cell(row=ROW_DAILY_TOTAL_FEMALE, column=current_col, value=f_female)
        
        # Combined: Row 75
        # =IF(G10="","",G43+G74)
        f_comb = f'=IF({col_letter}10="","",{col_letter}43+{col_letter}74)'
        sheet.cell(row=ROW_DAILY_TOTAL_COMBINED, column=current_col, value=f_comb)

        current_col += 1

    # Write Total School Days to AQ9
    sheet[COORD_TOTAL_DAYS] = valid_days_count
        
    # 3. Fill Students & Formulas
    _fill_student_section(sheet, data.get("students_male", []), ROW_START_MALE, ROW_END_MALE, date_col_map, holidays)
    _fill_student_section(sheet, data.get("students_female", []), ROW_START_FEMALE, ROW_END_FEMALE, date_col_map, holidays)
    
    wb.save(output_path)
    return output_path

def _fill_student_section(sheet, students, start_row, end_row, date_col_map, holidays):
    current_row = start_row
    
    for student in students:
        if current_row > end_row:
            break
            
        # Name at Column B (2)
        sheet.cell(row=current_row, column=2, value=student.get("name", ""))
        
        # Attendance
        attendance_records = student.get("attendance", {})
        
        for date_str, col_idx in date_col_map.items():
            if date_str in holidays:
                continue # Skip writing to holiday columns
                
            status = attendance_records.get(date_str, "PRESENT")
            val = "x" if status == "ABSENT" else ""
            sheet.cell(row=current_row, column=col_idx, value=val)
            
        # Inject Student Stats Formulas
        # Absences (AF / Col 32)
        # =IF(B13="","",COUNTIF(G13:AE13,"x") + COUNTIF(G13:AE13,"h")*0.5)
        col_af = openpyxl.utils.get_column_letter(COL_ABSENT)
        col_ag = openpyxl.utils.get_column_letter(COL_PRESENT)
        
        range_str = f"G{current_row}:AE{current_row}"
        f_abs = f'=IF(B{current_row}="","",COUNTIF({range_str},"x") + COUNTIF({range_str},"h")*0.5)'
        sheet.cell(row=current_row, column=COL_ABSENT, value=f_abs)
        
        # Presence (AG / Col 33)
        # =IF(B13="","",$AQ$9-AF13)
        f_pres = f'=IF(B{current_row}="","",$AQ$9-{col_af}{current_row})'
        sheet.cell(row=current_row, column=COL_PRESENT, value=f_pres)

        current_row += 1
